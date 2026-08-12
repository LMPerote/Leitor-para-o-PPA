"""Geração das planilhas Excel consolidadas (uma linha por documento).

Cada modelo de ficha gera seu próprio arquivo (``Indicadores.xlsx`` e
``Iniciativas.xlsx``), com as colunas definidas pelo modelo.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .modelos import COLUNA_ARQUIVO, Modelo

logger = logging.getLogger(__name__)

#: Limite do Excel por célula (32.767); usamos folga para o aviso de corte.
LIMITE_CELULA = 32_000
AVISO_CORTE = "... [texto truncado]"

LARGURA_MINIMA = 14
LARGURA_MAXIMA = 60
ALTURA_CABECALHO = 30
#: Nº de linhas amostradas para estimar a largura das colunas.
AMOSTRA_LARGURA = 500

_PREENCHIMENTO_CABECALHO = PatternFill("solid", fgColor="1F4E78")
_FONTE_CABECALHO = Font(bold=True, color="FFFFFF", size=11)
_ALINHAMENTO_CABECALHO = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALINHAMENTO_DADOS = Alignment(horizontal="left", vertical="top", wrap_text=True)
_BORDA = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def sanitizar(valor: object) -> str:
    """Deixa o texto seguro para o Excel.

    * remove caracteres de controle que invalidam o XML da planilha;
    * trunca no limite de caracteres por célula;
    * neutraliza textos iniciados por ``=`` (seriam lidos como fórmula).
    """
    texto = "" if valor is None else str(valor)
    texto = ILLEGAL_CHARACTERS_RE.sub("", texto)
    if len(texto) > LIMITE_CELULA:
        texto = texto[: LIMITE_CELULA - len(AVISO_CORTE)] + AVISO_CORTE
    if texto.startswith("="):
        texto = "'" + texto
    return texto


def montar_dataframe(registros: list[dict[str, str]], modelo: Modelo) -> pd.DataFrame:
    """Constrói o DataFrame de um modelo, na ordem de colunas que ele define."""
    colunas = modelo.colunas()
    quadro = pd.DataFrame(registros, columns=colunas, dtype="object")
    quadro = quadro.fillna("")
    for coluna in colunas:
        quadro[coluna] = quadro[coluna].map(sanitizar)
    if COLUNA_ARQUIVO in quadro.columns:
        quadro = quadro.sort_values(COLUNA_ARQUIVO, kind="stable").reset_index(drop=True)
    return quadro


def _formatar_aba(planilha, congelar: bool = True) -> None:
    """Negrito no cabeçalho, quebra de texto, alinhamento e largura das colunas."""
    total_colunas = planilha.max_column
    total_linhas = planilha.max_row

    for celula in planilha[1]:
        celula.font = _FONTE_CABECALHO
        celula.fill = _PREENCHIMENTO_CABECALHO
        celula.alignment = _ALINHAMENTO_CABECALHO
        celula.border = _BORDA
    planilha.row_dimensions[1].height = ALTURA_CABECALHO

    for linha in planilha.iter_rows(min_row=2, max_row=total_linhas):
        for celula in linha:
            celula.alignment = _ALINHAMENTO_DADOS
            celula.border = _BORDA

    # Largura estimada pelo maior conteúdo (limitada, pois há textos longos).
    # Em lotes grandes, amostrar as primeiras linhas basta e evita varrer
    # milhares de células só para calcular larguras.
    ultima_amostrada = min(total_linhas, AMOSTRA_LARGURA + 1)
    for indice in range(1, total_colunas + 1):
        letra = get_column_letter(indice)
        maior = len(str(planilha.cell(row=1, column=indice).value or ""))
        for linha in range(2, ultima_amostrada + 1):
            texto = str(planilha.cell(row=linha, column=indice).value or "")
            # Em textos com quebras, o que importa é a maior linha isolada.
            maior = max(maior, *(len(parte) for parte in texto.split("\n")))
        planilha.column_dimensions[letra].width = min(
            LARGURA_MAXIMA, max(LARGURA_MINIMA, maior + 2)
        )

    if congelar:
        planilha.freeze_panes = "C2"
        planilha.auto_filter.ref = (
            f"A1:{get_column_letter(total_colunas)}{max(total_linhas, 1)}"
        )


def salvar_excel(quadro: pd.DataFrame, destino: Path, modelo: Modelo) -> None:
    """Grava a planilha de um modelo: aba de dados + dicionário de campos."""
    destino.parent.mkdir(parents=True, exist_ok=True)
    dicionario = pd.DataFrame(modelo.dicionario())
    aba = destino.stem[:31] or "Dados"  # o Excel limita o nome da aba a 31

    with pd.ExcelWriter(destino, engine="openpyxl") as escritor:
        quadro.to_excel(escritor, sheet_name=aba, index=False)
        dicionario.to_excel(escritor, sheet_name="Dicionário", index=False)
        _formatar_aba(escritor.book[aba])
        _formatar_aba(escritor.book["Dicionário"], congelar=False)

    logger.info("Planilha gravada em %s (%d linha(s))", destino, len(quadro))


def salvar_csv(quadro: pd.DataFrame, destino: Path) -> None:
    """Exporta os mesmos dados em CSV (UTF-8 com BOM, separador ';')."""
    destino.parent.mkdir(parents=True, exist_ok=True)
    quadro.to_csv(destino, index=False, sep=";", encoding="utf-8-sig")
    logger.info("CSV gravado em %s", destino)
